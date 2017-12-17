# -*- coding: utf-8 -*-
"""
Created on Mon Jul 17 15:06:47 2017

@author: spran
"""

import requests
import os
import zipfile
import sqlite3
import glob
import csv
import xlsxwriter
import openpyxl
import numpy as np
from operator import itemgetter
import pandas as pd

# connecting to sqlite database and creating database 
conn = sqlite3.connect('medicare_hospital_compare.db')
c1 = conn.cursor()
staging_dir_name="staging"


def datafile_setup(staging_dir_name):
    """
    creating staging directory
    Un-compressing the zip file
    storing the csv files in the new directory created
    """
    url="https://data.medicare.gov/views/bg9k-emty/files/0a9879e0-3312-4719-a1db-39fd114890f1?content_type=application%2Fzip%3B%20charset%3Dbinary&filename=Hospital_Revised_Flatfiles.zip"   
    r = requests.get(url)

    # creating staging directory
    staging_dir_name="staging"
    if not os.path.isdir(staging_dir_name):
        os.mkdir(staging_dir_name)

    # extracting the zip file
    zip_file_name=os.path.join(staging_dir_name, "MedicareData.zip")
    zf=open(zip_file_name,"wb")
    zf.write(r.content)
    zf.close()
    z=zipfile.ZipFile(zip_file_name,'r')
    z.extractall(staging_dir_name)
    z.close()

    # ignoring the corrupt file as stated in the documentation
    if os.path.exists(staging_dir_name+"/FY2015_Percent_Change_in_Medicare_Payments.csv"):
        os.remove(staging_dir_name+"/FY2015_Percent_Change_in_Medicare_Payments.csv")

datafile_setup(staging_dir_name)
      
# SQLite Database Creation
cols = []
path = staging_dir_name+"/*.csv"
for fname in glob.glob(path):
    # changing the encoding of the csv file
    with open(fname,"rt",encoding='cp1252') as f:
        data = f.read()
    with open(fname,'wt',encoding='utf8') as f:
        f.write(data)
    # renaming the table names
    tbl_name = os.path.basename(fname).split('.')[0]
    if tbl_name[0].isalpha():
        tbl_name = tbl_name.replace(" ","_").replace("-","_").replace("%","pct").replace("/","_").lower()
    else:
        tbl_name = "t_" + str(tbl_name.replace(" ","_").replace("-","_").replace("%","pct").replace("/","_").lower())
    
    # creating the sqlite database all the csv files in the directory
    with open(fname, "rt", encoding="utf8") as f:
        reader = csv.reader(f)
        header = True
        for row in reader:
            if header:
                # collect the column names from the first row of the csv
                header = False
                # checking if the table already exists
                sql = "DROP TABLE IF EXISTS %s" % tbl_name
                c1.execute(sql)
                # renaming the columns of the all the tables
                for column in row:
                    if not column[0].isalpha():
                        column = "c_" + str(column.replace(" ","_").replace("-","_").replace("%","pct").replace("/","_").lower()) + " text"
                        cols.append(column)
                    else:
                        cols.append(column.replace(" ","_").replace("-","_").replace("%","pct").replace("/","_").lower()+ " text")
                # creating the table
                sql = "CREATE TABLE %s (%s)" % (tbl_name,
                        ", ".join(cols))
                c1.execute(sql)
                cols.clear()
                insertsql = "INSERT INTO %s VALUES (%s)" % (tbl_name,
                            ", ".join([ "?" for column in row ]))
                rowlen = len(row)
            else:
                # executing the insert query
                if len(row) == rowlen:
                   c1.execute(insertsql,tuple(row))
# commiting the changes to database
conn.commit()



#Download the ranking data file produced by in house system
url="http://kevincrook.com/utd/hospital_ranking_focus_states.xlsx"   
r = requests.get(url)
output = open('hospital_ranking_focus_states.xlsx','wb')
output.write(r.content)
output.close()

# creating of the Ranking workbook
ranking_wb = xlsxwriter.Workbook('hospital_ranking.xlsx')
# assigning the name for 1st worksheet
worksheet = ranking_wb.add_worksheet('Nationwide')
wb_row=0
wb_col=0
# creating the header for the worsheet
column_header=["Provider ID", "Hospital Name", "City", "State", "County"]
for col_name in column_header:
    worksheet.write(wb_row, wb_col, col_name)
    wb_col += 1

wb_row+=1
wb_col=0
# loading the ranking data file 
wb = openpyxl.load_workbook('hospital_ranking_focus_states.xlsx')

# assigning the name for 2nd worksheet
ws = wb.get_sheet_by_name('Hospital National Ranking')
# iterating through the worksheet
for row_index in ws.iter_rows('A2:A101'):
    for cell in row_index:
        # fetching the data based on the data from ranking worksheet
        c1.execute("SELECT provider_id, hospital_name, city, state, county_name FROM hospital_general_information WHERE provider_id = ?", (cell.value,))
        data=c1.fetchone()
        # storing the data into the new worksheet created
        worksheet.write(wb_row,wb_col,data[0])
        worksheet.write(wb_row,wb_col+1,data[1])
        worksheet.write(wb_row,wb_col+2,data[2])
        worksheet.write(wb_row,wb_col+3,data[3])
        worksheet.write(wb_row,wb_col+4,data[4])
        wb_row+=1

# creating worsheets for each states        
prov=[]
rank=[]
for names in ws.iter_rows(min_row=2, max_col=1, max_row=ws.max_row):
    for sd in names:
        prov.append(sd.value)

for names in ws.iter_rows(min_row=2, min_col=2, max_col=2, max_row=ws.max_row):
    for sd in names:
        rank.append(sd.value)
# creating dictionary for storing ranking and the provider id of the hospitals
data_dict = dict(zip(prov, rank))
# accessing the focus states data
ws = wb.get_sheet_by_name('Focus States')
tuple_name=tuple(ws['A2':'B11'])
for st in tuple_name:
    worksheet = ranking_wb.add_worksheet(st[0].value)
    row=0
    col=0
    # creating the header for the worsheet
    column_header=["Provider ID", "Hospital Name", "City", "State", "County"]
    for col_name in column_header:
        worksheet.write(row, col, col_name)
        col += 1
    # fetching data from database based on the states in the focus list
    c1.execute("SELECT provider_id, hospital_name, city, state, county_name FROM hospital_general_information WHERE state = ?", (st[1].value,))
    hosp_data=c1.fetchall()
    tuple_list=[list(x) for x in hosp_data]
    for tup in tuple_list:
        for key, value in data_dict.items():
            if tup[0] == key:
                tup.append(value)
    # sorting the list by ranking
    tuple_list.sort(key=itemgetter(5))   
    # writing the data in the list to the worksheet
    for count in range(1,101):
        for colm in range(0,5):
            worksheet.write(count,colm,tuple_list[count-1][colm])
        count+=1  

score_list=[]
# Creation of the Measures workbook
measure_wb = xlsxwriter.Workbook('measures_statistics.xlsx')
# assigning the name to 1st worksheet
measure_ws = measure_wb.add_worksheet('Nationwide')
row=0
col=0
# creating the header for the worsheet
column_header=["Measure ID", "Measure Name", "Minimum", "Maximum", "Average","Standard Deviation"]
for col_name in column_header:
    measure_ws.write(row, col, col_name)
    col += 1

row+=1
col=0
# retrieving distinct measure ids from the database
c1.execute("SELECT DISTINCT measure_id FROM timely_and_effective_care___hospital ORDER BY measure_id")
unique_id_list = [list(x) for x in c1.fetchall()]
#retriving the data for each unique measure id
for m_id in unique_id_list:
    c1.execute("SELECT measure_id, measure_name, score FROM timely_and_effective_care___hospital WHERE measure_id = ?", m_id)
    data_list = [list(x) for x in c1.fetchall()]
    for i in range(0,len(data_list)):
        if str(data_list[i][2]).isnumeric():
            score_list.append(int(data_list[i][2]))
            measure_ws.write(row,col,data_list[0][0])
            measure_ws.write(row,col+1,data_list[0][1])
    # writing the calculated data into the worksheet        
    if score_list:
        measure_ws.write(row,col+2,min(score_list))
        measure_ws.write(row,col+3,max(score_list))
        measure_ws.write(row,col+4,np.mean(score_list))
        measure_ws.write(row,col+5,np.std(score_list))
        score_list.clear()
    else:
        row-=1
    row+=1

distinct_m_id=[]
unique_id=[]
# retrieving distinct list of measure ids
c1.execute("SELECT DISTINCT measure_id FROM timely_and_effective_care___hospital")
distinct_m_id=c1.fetchall()
for m_id in distinct_m_id:
    unique_id.append(m_id[0])
# reading the focus list of states 
hosp_states = pd.read_excel('hospital_ranking_focus_states.xlsx', sheetname='Focus States', parse_cols="A:B" )

#creating dictionary of state names and the abbreviations
state_dict=hosp_states.set_index('State Name')['State Abbreviation'].to_dict()
for key, value in state_dict.items():
    measure_ws = measure_wb.add_worksheet(key)
    row=0
    col=0
    # creating the header for the worsheet
    column_header=["Measure ID", "Measure Name", "Minimum", "Maximum", "Average", "Standard Deviation"]
    for col_name in column_header:
        measure_ws.write(row, col, col_name)
        col += 1
    row+=1
    # retrieving data for numeric scores only for each state in the focus list 
    c1.execute("SELECT measure_id, measure_name FROM timely_and_effective_care___hospital WHERE abs(score) <> 0.0 AND state = ? GROUP BY measure_id", (value,))
    measure_data = c1.fetchall()
    list_tup = [list(x) for x in measure_data]
    #sorting based on measure id
    list_tup.sort(key=itemgetter(0))
    for tup in list_tup:
        for coln in range(0,2):
                measure_ws.write(row,coln,tup[coln])
        row+=1  
    # clearing the list    
    score_list.clear
    #retrieving the sorted measure ids from database
    c1.execute("SELECT DISTINCT measure_id FROM timely_and_effective_care___hospital WHERE state =? ORDER BY measure_id", (value,))
    unique_id_list = [list(x) for x in c1.fetchall()]
    row=1
    for u_id in unique_id_list:
        # retrieving data from the database
        c1.execute("SELECT measure_id, measure_name, score, state FROM timely_and_effective_care___hospital WHERE abs(score) <> 0.0 AND measure_id = ? ORDER BY measure_id", u_id)
        data_list = [list(x) for x in c1.fetchall()]
        for i in range(0,len(data_list)):
            if data_list[i][3] == value:
                score_list.append(int(data_list[i][2]))
        # writing the calculated data into the worksheet 
        if score_list:
            measure_ws.write(row,2,min(score_list))
            measure_ws.write(row,3,max(score_list))
            measure_ws.write(row,4,np.mean(score_list))
            measure_ws.write(row,5,np.std(score_list))
            score_list.clear()
            row+=1
            
# closing all the worbooks
ranking_wb.close()
measure_wb.close()

#closing the cursor 
c1.close()

# closing the sqlite connection
conn.close()